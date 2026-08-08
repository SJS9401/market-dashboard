"""
Manual CSV/XLSX data fetcher — historical 데이터 보강용.

Phase E-3: 1990 이전 사이클 보강. Yahoo/FRED 가 못 받는 시기 (1907 Panic, 1914-16 WW1,
1924-29 Roaring, 1978-80 Gold, 1979-80 Silver) 의 monthly 가격을 manual CSV/XLSX 에서 load.

지원 파일 형식 (3 종):
  1. Macrotrends 100Y CSV: "Date","Value" 2-col, Date = MM/DD/YYYY
  2. Shiller ie_data.xlsx: Data sheet, col 0=Date(decimal year), col 1=S&P Price
  3. (확장 가능): 다른 manual CSV format

표준 출력 (Yahoo/FRED 와 동일 schema):
  [[YYYYMMDD, O, H, L, C, V], ...]
  O=H=L=C=close, V=0 (monthly 데이터라 캔들 body 없음 → line chart 표시됨)

파일 위치: multi-asset/manual_data/{filename}
"""

import os
import sys
from pathlib import Path


HERE = Path(__file__).resolve().parent
MANUAL_DIR = HERE / 'manual_data'


def parse_macrotrends_csv(filename):
    """Macrotrends 'Date,Value' format → OHLC list.

    Format:
      ﻿"Date","Value"          (BOM utf-8-sig)
      "01/01/1915",19.25       (MM/DD/YYYY, value)

    Note: 1915-1959 의 경우 annual average 가 12개월 반복되는 경우 있음
    (Macrotrends 의 알려진 처리 방식). 1960+ 부터 진짜 monthly.
    """
    path = MANUAL_DIR / filename
    if not path.exists():
        return [], {'error': 'file not found: ' + str(path)}

    data = []
    try:
        with open(path, 'r', encoding='utf-8-sig') as f:
            lines = f.readlines()
    except Exception as e:
        return [], {'error': 'read failed: ' + str(e)}

    for line in lines[1:]:  # skip header
        line = line.strip()
        if not line:
            continue
        parts = line.split(',')
        if len(parts) < 2:
            continue
        date_raw = parts[0].strip().strip('"')
        val_raw = parts[1].strip().strip('"')
        try:
            mo, day, yr = date_raw.split('/')
            date_str = '{:04d}{:02d}{:02d}'.format(int(yr), int(mo), int(day))
            val = float(val_raw)
            data.append([date_str, val, val, val, val, 0])
        except Exception:
            continue

    # 정렬 (Macrotrends 가 ASC 로 주지만 안전 위해 sort)
    data.sort(key=lambda r: r[0])

    meta = {
        'source':   'manual-macrotrends',
        'file':     filename,
        'cadence':  'monthly',
        'start':    data[0][0] if data else None,
        'end':      data[-1][0] if data else None,
        'count':    len(data),
    }
    return data, meta


def parse_shiller_xlsx(filename, sheet='Data', date_col=0, price_col=1, data_start_row=9):
    """Shiller ie_data.xlsx Data sheet → OHLC list.

    Format:
      Row 7 (header): Date, P (S&P Price), D (Dividend), E (Earnings), ...
      Row 8+ (data): 1871.01, 4.44, 0.26, 0.4, ...

    Date decoding (decimal year):
      1871.01 → 1871 Jan
      1907.1  → 1907 Oct (= 1907.10, .0 truncated)
      1907.11 → 1907 Nov
      1907.12 → 1907 Dec
    """
    path = MANUAL_DIR / filename
    if not path.exists():
        return [], {'error': 'file not found: ' + str(path)}

    try:
        import openpyxl
    except ImportError:
        return [], {'error': 'openpyxl not installed (pip install openpyxl)'}

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        return [], {'error': 'open failed: ' + str(e)}

    if sheet not in wb.sheetnames:
        return [], {'error': 'sheet not found: ' + sheet + ' (available: ' + ','.join(wb.sheetnames) + ')'}

    ws = wb[sheet]
    data = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < data_start_row - 1:
            continue
        if row is None or len(row) <= max(date_col, price_col):
            continue
        d = row[date_col]
        p = row[price_col]
        if d is None or p is None:
            continue
        if not isinstance(d, (int, float)):
            continue
        if not isinstance(p, (int, float)):
            continue
        year = int(d)
        frac = d - year
        month = round(frac * 100)
        if month == 0:
            month = 1
        if month < 1 or month > 12:
            continue
        date_str = '{:04d}{:02d}01'.format(year, month)
        data.append([date_str, p, p, p, p, 0])

    data.sort(key=lambda r: r[0])

    meta = {
        'source':   'manual-shiller',
        'file':     filename,
        'sheet':    sheet,
        'cadence':  'monthly',
        'start':    data[0][0] if data else None,
        'end':      data[-1][0] if data else None,
        'count':    len(data),
    }
    return data, meta


def fetch_manual(spec):
    """
    Generic manual fetch dispatcher.

    spec: {'ticker': filename, 'parser': 'macrotrends' | 'shiller', ...}
    """
    parser = spec.get('parser', 'macrotrends')
    filename = spec.get('ticker', '')
    if not filename:
        return [], {'error': 'ticker (filename) missing in spec'}

    if parser == 'shiller':
        return parse_shiller_xlsx(filename)
    elif parser == 'macrotrends':
        return parse_macrotrends_csv(filename)
    else:
        return [], {'error': 'unknown parser: ' + parser}


if __name__ == '__main__':
    # CLI 테스트
    if len(sys.argv) < 2:
        print("Usage: python fetch_manual.py <filename> [parser]")
        print("  parsers: macrotrends (default) | shiller")
        sys.exit(1)
    filename = sys.argv[1]
    parser = sys.argv[2] if len(sys.argv) > 2 else 'macrotrends'
    data, meta = fetch_manual({'ticker': filename, 'parser': parser})
    print('Meta:', meta)
    if data:
        print('First:', data[0])
        print('Last: ', data[-1])
        print('Count:', len(data))
